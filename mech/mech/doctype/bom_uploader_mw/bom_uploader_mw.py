# Copyright (c) 2025, GreyCube Technologies and contributors
# For license information, please see license.txt

import frappe
from frappe import _, bold
from frappe.model.document import Document
from frappe.utils import cstr, get_link_to_form
from frappe.utils.xlsxutils import (
	build_xlsx_response,
	read_xlsx_file_from_attached_file,
	make_xlsx
)
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from copy import copy as xl_copy
import os
from io import BytesIO
from frappe.desk.utils import provide_binary_file
from erpnext.manufacturing.doctype.bom_creator.bom_creator import get_parent_row_no

TABLE_HEADERS = [
	"Row No",
	"Parent FG",
	"Sub Assembly Item",
	"Matched Item",	
	"Sr No",
	"User Input",
	"Description",
	"Length",
	"Width",
	"OD",
	"ID",
	"Thickness",
	"Material Type",
	"Qty",
	"GAD/MFG",
]

class BOMUploaderMW(Document):
	def validate(self):
		if self.import_excel:
			excel_data = self.read_excel()
			self.validate_imported_excel(excel_data)

		self.clear_table_data_if_not_attached_file()
		self.set_matched_item_in_bom_items()
		self.check_if_item_is_bought_out_or_restricted()
		self.calculate_raw_material_weight()

	# ----------------------------------------------------------------
	# Function Called on Click of Create BOM Creator button
	# ----------------------------------------------------------------
	@frappe.whitelist()
	def validate_conditions_and_create_bom_creator(self):
		self.check_if_all_matched_items_found_and_weigth_calculated()
		self.validate_duplicate_item()
		self.make_sub_assembly_items()
		self.make_bom_creator()

	# ----------------------------------------------------------------
	# This function was used to delete created sub assemblies on 
	# Cancel of BOM Uploader
	# ----------------------------------------------------------------
	# def on_cancel(self):
	# 	self.delete_all_sub_assembly_items()


	# ----------------------------------------------------------------
	# Function to Get Connected Sales Order Of Item 
	# ----------------------------------------------------------------
	@frappe.whitelist()
	def get_sales_order(self):
		so_item = frappe.db.get_value('Sales Order Item', {'item_code': self.dam_code}, 'parent')
		if not so_item:
			self.order_no = ""
			self.client = ""
			self.project = ""
			frappe.msgprint(_("For Item {0} No Sales Order Found.").format(self.dam_code), alert=1)
		else:
			customer, project = frappe.db.get_value('Sales Order', so_item, ['customer', 'project'])
			self.order_no = so_item
			self.client = customer
			self.project = project

	# ----------------------------------------------------------------
	# Function to Read Data Of Attached Excel 
	# ----------------------------------------------------------------
	@frappe.whitelist()
	def read_excel(self):
		file_doc = frappe.get_doc("File", {"file_url": self.import_excel})
		data = read_xlsx_file_from_attached_file(fcontent=file_doc.get_content())

		return data

	# ----------------------------------------------------------------
	# Function to Validate Attached Excel For Below Cases
	# - Validate File Name
	# - Check Data Exists
	# - Validate Mandatory Fields
	# - Validate Parent & Child Relationship
	# - Fill Item Details Table
	# ----------------------------------------------------------------
	def validate_imported_excel(self, excel_data):

		if len(self.bom_item_details_mw) < 1:
			# 1. Validate File Name
			file_name = frappe.db.get_value('File', { 'file_url': self.import_excel}, 'file_name')
			if file_name and self.name not in file_name:
				frappe.throw(_("Import Excel File name should be starts from {0}").format(self.name))

			# 2. Check Data Exists
			if len(excel_data) < 9:
				frappe.throw(_("Please add table data in excel"))

			# Convert raw data into Key-Value Format
			excel_table_data = self.get_excel_table_data(excel_data)

			# Build Item Level, 
			# Leaf items: no children — their own code never appears as parent_fg on any other row. 
			# They terminate a branch and represent an actual physical part (raw material or bought-out component), 
			# which is why they need Matched Item / material attributes.
			
			# Non-leaf items: have children — their own code appears as parent_fg on one or more other rows. 
			# They're structural/grouping nodes that become sub-assembly Items, not physical materials themselves.
			item_levels, leaf_items, non_leaf_items = self.build_item_tree_info(excel_table_data)
			table_header_col = excel_data[7]

			self.validate_excel_columns(table_header_col)
			self.check_in_excel_all_matrial_type_exists(excel_table_data)
			self.validate_mandatory_fields_and_matched_item_exist_in_excel(excel_table_data, leaf_items)
			self.validate_naming_and_sr_no_of_items(excel_table_data, leaf_items)
			self.fill_bom_item_details_table(excel_table_data, item_levels, leaf_items)

	# ----------------------------------------------------------------
	# Function to Convert Raw Data Into Key, Value Format Table Data
	# ----------------------------------------------------------------
	def get_excel_table_data(self, excel_data):
		data_len = len(excel_data)

		table_data = []
		for idx in range(8, data_len):
			row = excel_data[idx]
			# This condition will check for None values in all cells of Row
			if all(v is None for v in row) == False:
				# Remove Blank Row(Rows in which value is '')
				blank_row = True
				for a in row:
					if cstr(a).strip() != '':
						blank_row = False

				if blank_row == False:	
					table_data.append({
						"idx": idx + 1, 
						"row_no": row[0],
						"parent_fg": row[1],
						"sub_assembly_item": row[2],
						"matched_item": row[3],
						"sr_no": row[4],
						"user_input": row[5],
						"description": row[6],
						"length": row[7],
						"width": row[8],
						"od": row[9],
						"id": row[10],
						"thickness": row[11],
						"material_type": row[12],
						"qty": row[13],
						"gad_mfg": row[14]
					})
			else:
				pass

		return table_data

	# ----------------------------------------------------------------
	# Function to Generate Tree/Table Data From Excel Raw Data
	# ----------------------------------------------------------------
	def build_item_tree_info(self, excel_table_data):
		"""
		Positional (depth-first, pre-order) parent resolution. Assumes each
		row's children are entered immediately below it, before any sibling/
		uncle branch starts. This lets the same Sub Assembly Item code be
		reused as a placement under different parents at different levels
		(e.g. a shared sub-assembly used in more than one place in the tree)
		without the levels clashing.

		Leaf vs. non-leaf is decided by whether a row's own code is actually
		referenced as Parent FG by some other row — NOT by whether Sub
		Assembly Item is blank, since it can be filled in on every row
		(leaf and non-leaf alike).

		Returns:
			item_levels: dict {idx: level (int, 1-based)} — keyed by row idx,
				not by code, since the same code can appear at different
				levels in different placements.
			leaf_items: rows with no children (need Matched Item / weight calc)
			non_leaf_items: rows with children (structural sub-assembly rows)
		"""
		parent_codes_referenced = {
			row.get("parent_fg") for row in excel_table_data if row.get("parent_fg")
		}

		item_levels = {}
		leaf_items, non_leaf_items = [], []
		stack = []  # [(code, level), ...] currently open ancestors

		for row in excel_table_data:
			parent_fg = row.get("parent_fg")
			own_code = row.get("sub_assembly_item")
			idx = row.get("idx")
			is_non_leaf = bool(own_code) and own_code in parent_codes_referenced

			if parent_fg == self.dam_code:
				# Back to the FG root - close every previously open branch
				stack = []
				level = 1
			else:
				while stack and stack[-1][0] != parent_fg:
					stack.pop()
				if not stack:
					frappe.throw(
						_("In Excel Line No - {0}, Parent FG {1} is not correct. It should either be the Parent FG Item {2} or a Sub Assembly Item that was entered in an earlier row above this one.").format(
							idx, parent_fg, self.dam_code
						)
					)
				level = stack[-1][1] + 1

			item_levels[idx] = level

			if is_non_leaf:
				if any(code == own_code for code, _ in stack):
					frappe.throw(
						_("In Excel Line No - {0}, Item {1} cannot be a Sub Assembly of itself. Please check the Parent FG and Sub Assembly Item columns above this row.").format(idx, own_code)
					)
				stack.append((own_code, level))
				non_leaf_items.append(row)
			else:
				leaf_items.append(row)

		return item_levels, leaf_items, non_leaf_items

	# ----------------------------------------------------------------
	# Function to Validate Excel Columns
	# ----------------------------------------------------------------
	def validate_excel_columns(self, excel_column):
		a = excel_column
		b = TABLE_HEADERS
		is_equal = all(a == b for a, b in zip(a, b))
		if not is_equal:
			frappe.throw(_("In excel row 8 : Table Header Columns Must Be {0}").format(TABLE_HEADERS))

	# ----------------------------------------------------------------
	# Function to Check All Material Type Exists
	# ----------------------------------------------------------------	
	def check_in_excel_all_matrial_type_exists(self, excel_table_data):
		material_type_list = []
		for row in excel_table_data:
			material_type = row.get("material_type")
			if material_type and material_type not in material_type_list:
				material_type_list.append(material_type)

		if len(material_type_list) > 0:
			not_exists_mt = []
			for mt in material_type_list:
				if not frappe.db.exists("Material Type MW", mt):
					not_exists_mt.append(mt)

			if len(not_exists_mt) > 0:
				frappe.throw(_("Following material types are not exists: <br> {0}").format(",<br>".join((ele if ele != None else "") for ele in not_exists_mt)))

	# ----------------------------------------------------------------
	# Function to Validate Mandatory Fields In Excel
	# ----------------------------------------------------------------
	def validate_mandatory_fields_and_matched_item_exist_in_excel(self, excel_table_data, leaf_items):
		error_list = []
		for row in excel_table_data:
			excel_idx = row.get("idx")
			table_idx = row.get("row_no")

			not_exists_col = []
			if not row.get("row_no"):
				not_exists_col.append("<b>Row No</b>")
			if not row.get("parent_fg"):
				not_exists_col.append("<b>Parent FG</b>")
			if not row.get("sub_assembly_item"):
				not_exists_col.append("<b>Sub Assembly Item</b>")
			if not row.get("sr_no"):
				not_exists_col.append("<b>SR No</b>")
			if not row.get("description"):
				not_exists_col.append("<b>Description</b>")
			if not row.get("qty"):
				not_exists_col.append("<b>Qty</b>")
			if not row.get("gad_mfg"):
				not_exists_col.append("<b>GAD/MFG</b>")

			if row in leaf_items:
				# Validate Material Type For Leaf Item
				if not row.get("material_type"):
					not_exists_col.append("<b>Material Type</b>")

				# Validate Matched Item For Leaf Item
				if row.get("matched_item") and not frappe.db.exists("Item", row.get("matched_item")):
					not_exists_col.append("<b>Matched Item</b>")

				# If Material Type Exists Check For It's Attributes
				if row.get("material_type") and not row.get("matched_item"):
					mt = frappe.get_doc("Material Type MW", row.get("material_type"))
					if len(mt.attributes) > 0:
						for a in mt.attributes:
							excel_column_title = frappe.db.get_value(
								"Attribute MW", a.attribute, "excel_column_title"
							)
							lookup_key = (excel_column_title or "").strip().lower().replace(" ", "_")
							if not row.get(lookup_key):
								attr = "<b>" + (excel_column_title or a.attribute) + "</b>"
								not_exists_col.append(attr)

			if len(not_exists_col) > 0:
				if row.get("row_no"):
					msg = (
						"In Excel Line No - " + cstr(excel_idx) + ", Data Row No - " + cstr(table_idx) + " : " + (" and ".join((ele if ele != None else "") for ele in not_exists_col)))
				else:
					msg = ("In Excel Line No - " + cstr(excel_idx) + " : " + (" and ".join((ele if ele != None else "") for ele in not_exists_col )))

				error_list.append(msg)

		if len(error_list) > 0:
			frappe.throw(
				"Please Set Mandatory Field In Following Excel Columns: <br> {0}".format(
					",<br>".join((ele if ele != None else "") for ele in error_list)
				))

	# ----------------------------------------------------------------
	# Function to Validate Naming And Sr No Format Of Items
	# ----------------------------------------------------------------
	def validate_naming_and_sr_no_of_items(self, excel_table_data, leaf_items):
		naming_errors = []
		for item in excel_table_data:
			sr_no = item.get("sr_no")
			if not sr_no:
				continue

			if item in leaf_items:
				# Leaf Item's Sr No Must Be Alphanumeric
				if sr_no.isalpha():
					msg1 = (
						"In Excel Line No - {0}, Sr No should be alphanumeric not {1}"
					).format(item.get("idx"), sr_no)
					naming_errors.append(msg1)
			else:
				# Non Leaf Item's Sr No Must Be Alphabetic
				if not sr_no.isalpha():
					msg2 = (
						"In Excel Line No - {0}, Sr No should be alphabetic not {1}"
					).format(item.get("idx"), sr_no)
					naming_errors.append(msg2)

		if len(naming_errors) > 0:
			frappe.throw(
				"Please Correct Below Naming Errors: <br> {0}".format(
					",<br>".join((ele if ele != None else "") for ele in naming_errors)
				)
			)

	# ----------------------------------------------------------------
	# Function to Fill BOM Item Details Table
	# ----------------------------------------------------------------
	def fill_bom_item_details_table(self, excel_table_data, item_levels, leaf_items):
		if len(self.bom_item_details_mw) < 1:
			self.bom_item_details_mw = []
			for data in excel_table_data:
				item = self.append("bom_item_details_mw", {})
				item.row_no = data.get("row_no")
				item.parent_fg = data.get("parent_fg")
				item.sub_assembly_item = data.get("sub_assembly_item") or ""
				item.matched_item = data.get("matched_item") or ""
				item.sr_no = data.get("sr_no")
				item.user_input = data.get("user_input")
				item.description = data.get("description")
				item.length = data.get("length")
				item.width = data.get("width")
				item.od = data.get("od")
				item.id = data.get("id")
				item.thickness = data.get("thickness")
				item.material_type = data.get("material_type")
				item.qty = data.get("qty")
				item.gad_mfg = data.get("gad_mfg")

				item.level = item_levels.get(data.get("idx"))
				item.is_leaf_item = 1 if data in leaf_items else 0

				if item.is_leaf_item and item.matched_item and item.matched_item != "":
					item.status = "Match"

	# ----------------------------------------------------------------
	# Function to Clear Table Data If Excel File Is Not Attached
	# ----------------------------------------------------------------
	def clear_table_data_if_not_attached_file(self):
		if not self.import_excel:
			self.bom_item_details_mw = []

	# ----------------------------------------------------------------
	# Function to Set Matched Items
	# ----------------------------------------------------------------
	def set_matched_item_in_bom_items(self):
		if len(self.bom_item_details_mw) > 0:
			count = 0
			for item in self.bom_item_details_mw:
				if item.is_leaf_item:
					if item.matched_item:
						if not frappe.db.exists("Item", item.matched_item):
							frappe.throw(_("In Row No {0}, Matched Item {1} is not exists in Item Master.").format(item.row_no, item.matched_item))

					elif item.material_type and not item.matched_item:
						field_map = attributes_field_mapping()

						sub_assembly_item_group = frappe.db.get_single_value(
							"Mechwell Setting MW", "default_item_group_for_sub_assembly"
						)
						sql = "SELECT name FROM `tabItem` WHERE custom_material_type = '{0}' AND item_group !='{1}' ".format(
							item.material_type, sub_assembly_item_group
						)
						conditions = []
						near_by_value = {}

						attr_doc = frappe.get_doc("Material Type MW", item.material_type)

						is_sub_assembly_exists = False
						sub_assembly_keyword = ""
						if len(attr_doc.attributes) > 0:

							for att in attr_doc.attributes:
								att_map = frappe._dict(field_map[att.attribute])

								if att.attribute == "Sub Assembly Keyword":
									conditions.append(
										" ( %({})s like concat('%%', {}, '%%') ) ".format(
											att_map.field_name_in_bom_uploader,
											att_map.field_name_in_item_dt,
										)
									)
									is_sub_assembly_exists = True
									sub_assembly_keyword = (
										item.get(att_map.field_name_in_bom_uploader) or ""
									)
								elif att.match_type == ">=":
									conditions.append(
										" {field_name_in_item_dt} >=  %({field_name_in_bom_uploader})s ".format(
											**att_map
										)
									)
									max_value = frappe.db.sql_list(
										"SELECT min({field_name_in_item_dt}) FROM `tabItem` WHERE custom_material_type = '{0}' AND item_group !='{1}' AND {field_name_in_item_dt} >= %({field_name_in_bom_uploader})s".format(
											item.material_type,
											sub_assembly_item_group,
											**att_map,
										),
										item.as_dict(),
									)

									if max_value and max_value[0]:
										near_by_value[
											att_map.field_name_in_item_dt
										] = max_value[0]

								elif att.match_type == "<=":
									conditions.append(
										" {field_name_in_item_dt} <=  %({field_name_in_bom_uploader})s ".format(
											**att_map
										)
									)
									min_value = frappe.db.sql_list(
										"SELECT max({field_name_in_item_dt}) FROM `tabItem` WHERE custom_material_type = '{0}' AND item_group !='{1}' AND {field_name_in_item_dt} <= %({field_name_in_bom_uploader})s AND {field_name_in_item_dt} > 0".format(
											item.material_type,
											sub_assembly_item_group,
											**att_map,
										),
										item.as_dict(),
									)
									if min_value and min_value[0]:
										near_by_value[
											att_map.field_name_in_item_dt
										] = min_value[0]

								elif att.match_type == "==":
									conditions.append(
										" ( {} = %({})s ) ".format(
											att_map.field_name_in_item_dt,
											att_map.field_name_in_bom_uploader,
										)
									)

							if conditions:
								sql = sql + " AND " + " AND ".join(conditions)

							matched_items = frappe.db.sql(sql, item.as_dict(), pluck="name")

							final_matched_items = []
							if len(matched_items) > 0:
								if near_by_value:

									for i in matched_items:
										item_doc = frappe.get_doc("Item", i)

										for key, value in near_by_value.items():
											if (
												key in item_doc.as_dict()
												and item_doc.get(key) == value
											):
												if item_doc.name not in final_matched_items:
													final_matched_items.append(item_doc.name)
													continue

								if len(final_matched_items) > 0:
									item.matched_item_list = ",".join(final_matched_items)
									if len(final_matched_items) == 1:
										item.matched_item_list = final_matched_items[0]
										item.matched_item = final_matched_items[0]

									else:
										if is_sub_assembly_exists == True:
											exact_matched_items = (
												check_exact_matched_sub_assembly_item(
													matched_items, sub_assembly_keyword
												)
											)
											if len(exact_matched_items) == 1:
												item.matched_item = exact_matched_items[0]
												item.matched_item_list = exact_matched_items[0]
											elif len(exact_matched_items) > 1:
												item.matched_item_list = ",".join(
													exact_matched_items
												)
												item.status = "Multi Match"
											else:
												item.status = "Multi Match"
										else:
											item.status = "Multi Match"
								else:
									item.matched_item_list = ",".join(matched_items)
									if len(matched_items) == 1:
										item.matched_item_list = matched_items[0]
										item.matched_item = matched_items[0]
									else:
										if is_sub_assembly_exists == True:
											exact_matched_items = (
												check_exact_matched_sub_assembly_item(
													matched_items, sub_assembly_keyword
												)
											)
											if len(exact_matched_items) == 1:
												item.matched_item = exact_matched_items[0]
												item.matched_item_list = exact_matched_items[0]
											elif len(exact_matched_items) > 1:
												item.matched_item_list = ",".join(
													exact_matched_items
												)
												item.status = "Multi Match"
											else:
												item.status = "Multi Match"
										else:
											item.status = "Multi Match"

								if item.matched_item:
									item_group, custom_wmf = frappe.db.get_value(
										"Item", item.matched_item, ["item_group", "custom_wmf"]
									)
									item.matched_item_group = item_group
									item.item_wmf = custom_wmf
									item.status = "Match"

							else:
								item.status = "Not Found"
						else:
							item.status = "Not Found"
					

				count += 1
				frappe.publish_progress(
					count / len(self.bom_item_details_mw) * 100,
					title="Finding Matching Items",
					description="",
				)
		
	# ----------------------------------------------------------------
	# Function to Check Item is Boughtout / Restricted
	# ----------------------------------------------------------------
	def check_if_item_is_bought_out_or_restricted(self):
		mech_setting = frappe.get_doc('Mechwell Setting MW')
		restricted_item_groups = []
		if len(mech_setting.restricted_item_groups) > 0:
			for d in mech_setting.restricted_item_groups:
				restricted_item_groups.append(d.restricted_item_group)

		default_item_group_for_bought_out = mech_setting.default_item_group_for_bought_out
		if not default_item_group_for_bought_out:
			frappe.throw(_("Please set Default Item Group for Bought Out In Mechwell Settings Doctype."))

		if len(self.bom_item_details_mw) > 0:
			for item in self.bom_item_details_mw:
				if item.is_leaf_item and item.matched_item:
					item_group = frappe.db.get_value('Item', item.matched_item, 'item_group')
					item.create_subassembly_item = "Yes"

					is_bought_out = check_if_item_is_bought_out_or_restricted(default_item_group_for_bought_out, item_group)
					is_restricted_item_group = False
					for rig in restricted_item_groups:
						if check_if_item_is_bought_out_or_restricted(rig, item_group):
							is_restricted_item_group = True
							break
					if is_bought_out:
						item.is_bought_out = 'Yes'
						item.create_subassembly_item = "No"
					else:
						item.is_bought_out = 'No'
						
					if is_restricted_item_group and item.is_bought_out == "No":
						item.create_subassembly_item = "No"

	# ----------------------------------------------------------------
	# Function to Check Duplicate Matched Item Under Same Parent
	# ----------------------------------------------------------------
	def validate_duplicate_item(self):
		item_map = {}
		for row in self.bom_item_details_mw:
			if not (row.is_leaf_item and row.matched_item and row.create_subassembly_item == "No"):
				continue

			key = (row.parent_fg, row.matched_item)
			if key in item_map:
				frappe.throw(
					_("Item {0} added multiple times under the same parent item {1} at Row No {2} and {3}.<br>{4}").format(
						bold(row.matched_item),
						bold(row.parent_fg),
						item_map[key],
						row.row_no,
						_("This happens when the Sub Assembly Item's Raw Material is Bought Out or belongs to a Restricted Item Group, which forces it to be added directly under the parent instead of creating a Sub Assembly."),
					),
					title=_("Duplicate Item Under Same Parent"),
				)
			else:
				item_map[key] = row.row_no

	# ----------------------------------------------------------------
	# Function to Calculates Each Item's Weight & Total Weight
	# ----------------------------------------------------------------
	def calculate_raw_material_weight(self):
		total_raw_weight = 0
		for item in self.bom_item_details_mw:
			if item.is_leaf_item and item.matched_item:
				if item.is_bought_out == "Yes":
					item.raw_material_weight = item.qty * (item.item_wmf or 0) 
				else:
					ig, wmf = frappe.db.get_value("Item", item.matched_item, ["item_group", "custom_wmf"])
					item_group = frappe.get_doc('Item Group', ig)
					formula = item_group.custom_raw_material_weight_formula
					formula_params = {
						'L': item.length or 0,
						'W': item.width or 0,
						'T': item.thickness or 0,
						'D': item_group.custom_density or 0,
						'OD' : item.od or 0,
						'ID' : item.id or 0,
						'WPM' : wmf or 0,
						'PPW' : wmf or 0,
						'TP' : item.qty or 0,
						'π': 3.14
					}
					if item_group.custom_is_od_formula_exists == 1 and item.od:
						formula = item_group.custom_od_based_weight_formula or None
					else:
						formula = item_group.custom_raw_material_weight_formula or None

					if not formula:
						frappe.throw(_("Please set Raw Material Weight Formula in Item Group <b>{0}</b>").format(get_link_to_form("Item Group", item_group.name)))
						
					total_weight = frappe.safe_eval(formula.strip(), None, formula_params)

					item.raw_material_weight = total_weight or 0

				total_raw_weight = total_raw_weight + (item.raw_material_weight or 0)
				self.total_weight = total_raw_weight

	# ----------------------------------------------------------------------------
	# Function to Recompute Matched Item Backfill & Weights For The Tree Preview
	# ----------------------------------------------------------------------------
	@frappe.whitelist()
	def recalculate_bom_weights(self):
		for item in self.bom_item_details_mw:
			if item.is_leaf_item and item.matched_item and (not item.matched_item_group or not item.item_wmf):
				item_group, custom_wmf = frappe.db.get_value("Item", item.matched_item, ["item_group", "custom_wmf"])
				item.matched_item_group = item_group
				item.item_wmf = custom_wmf
				item.status = "Match"
		self.check_if_item_is_bought_out_or_restricted()
		self.calculate_raw_material_weight()

	# ----------------------------------------------------------------------------
	# Function to Check All Match Items & Weights is Calculated Before Submit
	# ----------------------------------------------------------------------------
	def check_if_all_matched_items_found_and_weigth_calculated(self):
		if len(self.bom_item_details_mw) > 0:
			item_not_found = []
			weight_not_calculated = []
			for row in self.bom_item_details_mw:
				if row.is_leaf_item:
					if not row.matched_item:
						item_not_found.append(cstr(row.idx))
					if not row.raw_material_weight or row.raw_material_weight == 0:
						print("Row No: ", row.idx, "Weight: ", row.raw_material_weight)
						weight_not_calculated.append(cstr(row.idx))
				else:
					continue
			
			if len(item_not_found) > 0:
				frappe.throw(_("For Below Row Numbers Match Item Not Found.<br> <b>{0}</b>").format(", ".join((ele if ele != None else "") for ele in item_not_found)))

			if len(weight_not_calculated) > 0:
				frappe.throw(_("For Below Row Numbers Weight Is Not Calculated.<br> <b>{0}</b>").format(", ".join((ele if ele != None else "") for ele in weight_not_calculated)))

	# ----------------------------------------------------------------------------
	# Function to Create Sub Assembly Items Before Submit
	# ----------------------------------------------------------------------------
	def make_sub_assembly_items(self):
		if len(self.bom_item_details_mw) > 0:
			sub_assembly_item_group = frappe.db.get_single_value('Mechwell Setting MW', 'default_item_group_for_sub_assembly')
			for row in self.bom_item_details_mw:
				if row.create_subassembly_item != "No" and not frappe.db.exists("Item", row.sub_assembly_item):
					new_item = frappe.new_doc("Item")
					new_item.item_code = row.sub_assembly_item
					new_item.item_name = row.sub_assembly_item + ((" " + row.user_input) if row.user_input else "")
					new_item.item_group = sub_assembly_item_group
					new_item.description = row.description
					new_item.custom_length = row.length
					new_item.custom_width = row.width
					new_item.custom_outer_diameter = row.od
					new_item.custom_inner_diameter = row.id
					new_item.custom_thickness = row.thickness

					new_item.save(ignore_permissions=True)

					row.sub_assembly_item = new_item.name

	# ----------------------------------------------------------------------------
	# Function to Create BOM Creater On Submit Of BOM Uploader
	# ----------------------------------------------------------------------------
	def make_bom_creator(self):
		if len(self.bom_item_details_mw) > 0:
			bom = frappe.new_doc("BOM Creator")
			bom.__newname = self.name
			bom.name = self.name
			bom.item_code = self.dam_code
			bom.qty = 1
			bom.custom_bom_uploader_ref = self.name
			bom.project = self.project

			# Positional (depth-first, pre-order) parent resolution — mirrors
			# build_item_tree_info. A plain DB lookup by Sub Assembly Item code
			# can't disambiguate a reused code (the same sub-assembly placed
			# under multiple different parents), so we walk the rows in order
			# and track which ancestor is currently open instead.
			stack = []  # [(sub_assembly_item, bom.items row idx), ...] currently open ancestors

			for row in self.bom_item_details_mw:
				item = bom.append("items", {})

				if row.parent_fg == self.dam_code:
					stack = []
					parent_idx = None
				else:
					while stack and stack[-1][0] != row.parent_fg:
						stack.pop()
					if not stack:
						frappe.throw(
							_("Row No {0}: Parent FG {1} not found while creating BOM Creator").format(row.row_no, row.parent_fg)
						)
					parent_idx = stack[-1][1]

				item.fg_item = row.parent_fg
				item.qty = row.qty
				item.custom_sr_no = row.sr_no
				item.parent_row_no = parent_idx
				if row.gad_mfg == "GAD":
					item.allow_alternative_item = 0
				else:
					item.allow_alternative_item = 1

				if row.create_subassembly_item == "No":
					item.is_expandable = 0
					item.item_code = row.matched_item
					if row.is_bought_out == "Yes":
						item.custom_is_bought_out = row.is_bought_out
						item.qty = row.raw_material_weight
				else:
					item.item_code = row.sub_assembly_item
					item.is_expandable = 1
					if row.matched_item and row.is_leaf_item:
						raw_item = bom.append("items", {})
						raw_item.item_code = row.matched_item
						raw_item.fg_item = item.item_code
						raw_item.qty = row.raw_material_weight
						raw_item.parent_row_no = item.idx
						raw_item.uom = frappe.db.get_value("Item", row.matched_item, "stock_uom")
						if row.gad_mfg == "GAD":
							raw_item.allow_alternative_item = 0
						else:
							raw_item.allow_alternative_item = 1
				item.uom = frappe.db.get_value("Item", item.item_code, "stock_uom")
				if not row.is_leaf_item:
					stack.append((row.sub_assembly_item, item.idx))

			bom.save(ignore_permissions=True)
			self.db_set("bom_creator_ref", bom.name)
			frappe.msgprint(_("BOM Creator {0} is created.").format(get_link_to_form("BOM Creator", bom.name)))

	# ----------------------------------------------------------------------------
	# Function to Delete Sub Assembly Items On Delete Of BOM Uploader
	# ----------------------------------------------------------------------------
	def delete_all_sub_assembly_items(self):
		bom_creator = frappe.db.get_value("BOM Creator", {"custom_bom_uploader_ref": self.name}, "name")
		if len(self.bom_item_details_mw) > 0 and not bom_creator:
			for row in self.bom_item_details_mw:
				if row.sub_assembly_item:
					sub_assembly_item = frappe.get_doc("Item", row.sub_assembly_item)
					row.sub_assembly_item = ""
					sub_assembly_item.delete()
				
			frappe.msgprint("Sub Assembly Items Are Deleted", alert=1)

# ----------------------------------------------------------------------------
# HELPER FUNCTIONS
# ----------------------------------------------------------------------------
def check_if_item_is_bought_out_or_restricted(default_item_group_for_bought_out_or_restricted, item_group):
	if item_group == default_item_group_for_bought_out_or_restricted:
		return True
	elif item_group != default_item_group_for_bought_out_or_restricted:
		parent_item_group = frappe.db.get_value('Item Group', item_group, 'parent_item_group')
		if parent_item_group == default_item_group_for_bought_out_or_restricted:
			return True
		elif not parent_item_group:
			return False
		elif parent_item_group and parent_item_group != default_item_group_for_bought_out_or_restricted:
			check_if_item_is_bought_out_or_restricted(default_item_group_for_bought_out_or_restricted, parent_item_group)
	
def attributes_field_mapping():
	attribute_mw = frappe.db.get_all('Attribute MW', fields=['name', 'field_name_in_item_dt', 'field_name_in_bom_uploader'])
	field_map = {}
	for d in attribute_mw:
		field_map[d.name] = d     

	return field_map

def check_exact_matched_sub_assembly_item(item_list, sub_assembly_keyword):
		exact_matched_items = []
		if len(item_list) > 0:
			for item in item_list:
				item_keyword = frappe.db.get_value('Item', item, 'custom_sub_assembly_keyword')
				if item_keyword.strip().casefold() == sub_assembly_keyword.strip().casefold():
					exact_matched_items.append(item)

		return exact_matched_items	


# ----------------------------------------------------------------
# Function to download Formatted Excel Template 
# ----------------------------------------------------------------
@frappe.whitelist()
def download_formatted_excel(name=None, add_table=False):
	doc = frappe.get_doc("BOM Uploader MW", name) if name else None
	print("Downloading formatted Excel file...")

	workbook = Workbook()
	sheet = workbook.active

	rows_data = [
		["", "", "Dam code", doc.dam_code],
		["", "", "Order No", doc.order_no],
		["", "", "Client", doc.client],
		["", "", "Project", doc.project],
		["", "", "Wt(kg)", doc.total_weight],
		["Instruction : Please input data from row no 9. Donot put blank rows while data input"],
		[],
		TABLE_HEADERS,
	]

	for row in rows_data:
		sheet.append(row)

	if add_table:
		if len(doc.bom_item_details_mw) > 0:
			for row in doc.bom_item_details_mw:
				sheet.append([
					row.row_no,
					row.parent_fg,
					row.sub_assembly_item,
					row.matched_item,
					row.sr_no,
					row.user_input,
					row.description,
					row.length,
					row.width,
					row.od,
					row.id,
					row.thickness,
					row.material_type,
					row.qty,
					row.gad_mfg,
				])

	sheet.column_dimensions['A'].width = 10
	sheet.column_dimensions['B'].width = 12
	sheet.column_dimensions['C'].width = 20
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 10
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50
	sheet.column_dimensions['H'].width = 10
	sheet.column_dimensions['I'].width = 10
	sheet.column_dimensions['J'].width = 10
	sheet.column_dimensions['K'].width = 10
	sheet.column_dimensions['L'].width = 10
	sheet.column_dimensions['M'].width = 20
	sheet.column_dimensions['N'].width = 10
	sheet.column_dimensions['O'].width = 10

	bg_fill = PatternFill(fill_type='solid', start_color='FF474C', end_color='FF474C')

	cells_to_style = ['A8', 'B8', 'C8', 'E8', 'G8', 'M8', 'N8', 'O8']
	for cell_coord in cells_to_style:
		cell = sheet[cell_coord]
		cell.fill = bg_fill

	xlsx_file = BytesIO()
	workbook.save(xlsx_file)

	provide_binary_file(doc.name, 'xlsx', xlsx_file.getvalue())
