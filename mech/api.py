import frappe
from frappe import _
from frappe.utils import cint, cstr
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file
import re
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import os
from io import BytesIO
from frappe.desk.utils import provide_binary_file



def get_operation_table_header():
	TABLE_HEADERS = [ "Row No", "Parent FG", "BOM Item Code", "Sr No", "Description", "Length", "Width", 
				 "OD", "ID", "Thickness", "Material Type", "Qty", "GAD/MFG", "Is Bought Out"]
	
	operation_list = frappe.db.sql_list("SELECT name FROM tabOperation ORDER BY if(custom_priority='0',1,0) ASC,custom_priority ASC")
	if len(operation_list) > 0 :
		for operation in operation_list:
			if operation not in TABLE_HEADERS:
				TABLE_HEADERS.append(operation)
	
	return TABLE_HEADERS

@frappe.whitelist()
def download_operation_formatted_excel(bom_uploader=None, name=None):
	doc = frappe.get_doc("BOM Uploader MW", bom_uploader)
	print("Downloading Operation formatted Excel file...")
	
	workbook = Workbook()
	sheet = workbook.active

	table_headers = get_operation_table_header()
	
	rows_data = [
		["", "", "Dam code", doc.dam_code],
		["", "", "Order No", doc.order_no],
		["", "", "Client", doc.client],
		["", "", "Project", doc.project],
		["", "", "Wt(kg)", doc.total_weight],
		["Instruction : \nPlease input data from row no 9. Do not put blank rows while data input \n Is Bought column is read only \n Operation time to be input in mins"],
		[],
		table_headers,
	]

	if len(doc.bom_item_details_mw) > 0:

		for row in doc.bom_item_details_mw:
			row_details = [ row.row_no, row.parent_fg, row.sub_assembly_item or '', row.sr_no, row.description, '' if row.length == 0 else row.length, '' if row.width == 0 else row.width, 
				  '' if row.od == 0 else row.od ,'' if row.id == 0 else row.id, '' if row.thickness == 0 else row.thickness, row.material_type or '', row.qty, row.gad_mfg]
			
			if row.is_bought_out == "Yes":
				row_details[2] = row.matched_item
				row_details.append(row.is_bought_out)
			else:
				pass

			# print(row_details, "=============row_details===========")
			rows_data.append(row_details)

	for row in rows_data:
		sheet.append(row)

	sheet.column_dimensions['A'].width = 10
	sheet.column_dimensions['B'].width = 12
	sheet.column_dimensions['C'].width = 15
	sheet.column_dimensions['D'].width = 10
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 10
	sheet.column_dimensions['G'].width = 10
	sheet.column_dimensions['H'].width = 10
	sheet.column_dimensions['I'].width = 10
	sheet.column_dimensions['J'].width = 10
	sheet.column_dimensions['K'].width = 15
	sheet.column_dimensions['L'].width = 10
	sheet.column_dimensions['M'].width = 15
	sheet.column_dimensions['N'].width = 15

	for column_cells in sheet.columns:
		sheet.column_dimensions[column_cells[0].column_letter].auto_size = True

	bg_fill = PatternFill(fill_type='solid', start_color='FF474C', end_color='FF474C')

	cells_to_style = ['A8', 'B8', 'D8','E8', 'K8', 'L8', 'M8']
	for cell_coord in cells_to_style:
		cell = sheet[cell_coord]
		cell.fill = bg_fill
	
	for cell in sheet["E:E"]:
		cell.alignment=Alignment(wrap_text=True)

	xlsx_file = BytesIO()
	workbook.save(xlsx_file)

	provide_binary_file(name, 'xlsx', xlsx_file.getvalue())

def validate_operation_excel(self, method):
	if self.custom_attach_operation_data and len(self.items) < 1:
		validate_excel_file_name(self.name, self.custom_attach_operation_data)

		excel_data = read_excel_data(self.custom_attach_operation_data)
		bom_uploader_sub_assembly_items = frappe.db.sql_list("SELECT IF(is_bought_out='Yes', matched_item, sub_assembly_item ) FROM `tabBOM Item Details MW` WHERE parent='{0}'".format(self.custom_bom_uploader_ref))

		# Table Header Validation
		if len(excel_data) > 0:
			excel_table_header_col = excel_data[7]
			table_headers = get_operation_table_header()
			validate_excel_columns(excel_table_header_col, table_headers)

			# Validate Sub Assembly & Operation Values
			item_error = []
			operation_error = []
			for row_data in range(8, len(excel_data)):
				row = excel_data[row_data]
				sub_assembly_item = row[2]

				for op in range(14, len(table_headers)):
					if row[op] and isinstance(row[op], (float, int)) == False:
						msg2 = ("In Excel Line No - " + cstr(cint(row)+1) + " Data Row No - " + cstr(row[0]) + " Column - " + table_headers[op] + ": " + row[op])
						operation_error.append(msg2)

				if sub_assembly_item not in bom_uploader_sub_assembly_items:
					msg1 = ("In Excel Line No - " + cstr(cint(row)+1) + " Data Row No - " + cstr(row[0]) + " Sub Assembly Item " + sub_assembly_item)
					item_error.append(msg1)

			# print(operation_error, "========operation_error====")
			# print(item_error, "============item_error===")
			if len(item_error) > 0 and len(operation_error) > 0:
				frappe.throw(_("Below Sub Assembly Items are not Found in BOM Uploader {0} : <br> {1} <br><br> Below Operation's time must be in numeric value: <br>{2}").format(self.custom_bom_uploader_ref, ",<br>".join((ele if ele != None else "") for ele in item_error), ",<br>".join((ele if ele != None else "") for ele in operation_error) ))
			elif len(item_error) > 0:
				frappe.throw(_("Below Sub Assembly Items are not Found in BOM Uploader {0} : <br> {1}").format(self.custom_bom_uploader_ref, ",<br>".join((ele if ele != None else "") for ele in item_error) ))
			elif len(operation_error) > 0:
				frappe.throw(_("Below Operation's time must be in numeric value: <br> {0}").format(",<br>".join((ele if ele != None else "") for ele in operation_error)))

def fill_operation_table_from_excel_data(self, method):
	print("==============fill_operation_table_from_excel_data=============")
	if self.custom_attach_operation_data and len(self.custom_operations_details) == 0:
		excel_data = read_excel_data(self.custom_attach_operation_data)
		table_headers = excel_data[7]

		data_len = len(excel_data)
		table_row_len = len(table_headers)

		table_data = []
		for idx in range(8, data_len):
			row = excel_data[idx]
			if all(v is None for v in row) == False:

				##### remove blank row for data #####
				blank_row = True
				for a in row:
					if cstr(a).strip() != '':
						blank_row = False

				if blank_row == False:	
					table_row_data = []

					table_row_data.append({
						"bom_item_code": row[2],
					})
				
					for operation_idx in range(14, table_row_len):
						table_row_data.append({
							"operation": table_headers[operation_idx],
							"value" : row[operation_idx] or 0
						})
					table_data.append(table_row_data)

		# print(table_data)

		if len(table_data) > 0:
			for row in table_data:
				for op in range(1, len(row)):
					if row[op].get("value") and cint(row[op].get("value")) > 0:
						is_bought_out = frappe.db.get_value("BOM Creator Item", {"parent": self.name, "item_code": row[0].get("bom_item_code")}, "custom_is_bought_out")
						# print(row[0].get("bom_item_code"), "=======", is_bought_out, "=====is_bought_out===")
						if is_bought_out != "Yes":
							self.append("custom_operations_details",{
								"item": row[0].get("bom_item_code"),
								"operation": row[op].get("operation"),
								"operation_time": row[op].get("value")
							})

def clear_operation_table_if_not_attached_excel(self, method):
	if not self.custom_attach_operation_data:
		self.custom_operations_details = []

def validate_excel_file_name(doc_name, attached_excel):
	file_name = frappe.db.get_value('File', { 'file_url': attached_excel}, 'file_name')
	if file_name and doc_name not in file_name:
		frappe.throw(_("Import Excel File name should be starts from {0}").format(doc_name))

def read_excel_data(attached_excel):
	file_doc = frappe.get_doc("File", {"file_url": attached_excel})
	excel_data = read_xlsx_file_from_attached_file(fcontent=file_doc.get_content())

	return excel_data

def validate_excel_columns(excel_column, table_headers):
		a = excel_column
		b = table_headers
		is_equal = all(a == b for a, b in zip(a, b))
		if not is_equal:
			frappe.throw(_("In excel row 8 : Table Header Columns Must Be {0}").format(
					table_headers))

def check_workstation_exist_for_all_operation(self, method):
	if len(self.custom_operations_details) > 0:
		unique_op = []
		for op in self.custom_operations_details:
			if op.operation not in unique_op:
				unique_op.append(op.operation)
		
		workstation_not_found = []
		for d in unique_op:
			workstation = frappe.db.get_value("Operation", d, "workstation")
			if not workstation:
				workstation_not_found.append(d)

		if len(workstation_not_found) > 0:
			frappe.throw(_("Please Set WorkStation For Below Operations: <br> {0}").format(",<br>".join((ele if ele != None else "") for ele in workstation_not_found)))
			

def add_operation_from_bom_creator(self, method):
	if self.bom_creator and self.bom_creator_item:
		# print(self.bom_creator , "====================")
		allow_alternative_item = frappe.db.get_value("BOM Creator Item", self.bom_creator_item, "allow_alternative_item")
		if allow_alternative_item == 1:
			self.allow_alternative_item = 1
			self.custom_gad_mfg = "MFG"
		else:
			self.allow_alternative_item = 0
			self.custom_gad_mfg = "GAD"
			
		bom_uploader = frappe.db.get_value("BOM Creator", self.bom_creator, "name")
		if bom_uploader:
			bom_creator = frappe.get_doc("BOM Creator", self.bom_creator)
			if len(bom_creator.custom_operations_details) > 0 and len(self.operations) == 0:
				for op in bom_creator.custom_operations_details:
					
					if op.item == self.item:
						# print(op.item, "=============op.item======")
						if self.with_operations == 0:
							self.with_operations = 1

						workstation = frappe.db.get_value("Operation", op.operation, "workstation")

						self.append("operations", {
							"operation": op.operation,
							"time_in_mins": op.operation_time,
							"workstation": workstation
						})